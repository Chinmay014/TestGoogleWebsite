import inspect
import logging
import csv
from openpyxl import workbook,load_workbook
class custLogger:
    def customlogger(self, loglevel = logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        fh = logging.FileHandler("reports/Debug.log")
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger
    
    def read_data_from_excel(file_name,sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_count = sh.max_row
        column_count = sh.max_column

        for i in range(2,row_count+1):
            row = []
            for j in range(1,column_count+1):
                row.append(sh.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist
            
    def read_data_from_csv(file_name):
        datalist=[]
        reader = csv.reader(open(file_name,"r"))
        # skip reader
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist